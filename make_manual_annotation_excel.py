from pathlib import Path
import pandas as pd

# ------------------------------------------------------------
# Paths
# ------------------------------------------------------------
BASE_DIR = Path(r"C:\Users\Max Stauffer\BreakingRules")

INPUT_CSV = BASE_DIR / "sample_50_for_llm_prompt_calibration.csv"

OUTPUT_CSV = BASE_DIR / "manual_annotation_50_surgical_tech.csv"
OUTPUT_XLSX = BASE_DIR / "manual_annotation_50_surgical_tech.xlsx"

# ------------------------------------------------------------
# Read sample file
# ------------------------------------------------------------
if not INPUT_CSV.exists():
    raise FileNotFoundError(f"Input file not found: {INPUT_CSV}")

df = pd.read_csv(INPUT_CSV, encoding="utf-8-sig", low_memory=False)

# ------------------------------------------------------------
# Required source columns
# ------------------------------------------------------------
required_source_cols = ["pmid", "year", "journal", "title", "abstract"]

missing = [c for c in required_source_cols if c not in df.columns]
if missing:
    raise ValueError(f"Missing required columns in input CSV: {missing}")

# ------------------------------------------------------------
# Build manual annotation table
# ------------------------------------------------------------
out = df[required_source_cols].copy()

out["manual_decision"] = ""              # include / reject
out["reason"] = ""                       # e.g. not visceral, no technical solution, include: robotic visceral surgery
out["manual_surgical_problem"] = ""      # only if include
out["manual_technical_solution"] = ""    # only if include
out["augmentation_bridge"] = ""          # optional only

# Keep exactly 50 rows if file contains more
out = out.head(50)

# ------------------------------------------------------------
# Save CSV
# ------------------------------------------------------------
out.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")

# ------------------------------------------------------------
# Save clean Excel
# ------------------------------------------------------------
with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
    out.to_excel(writer, index=False, sheet_name="manual_annotation")

    ws = writer.book["manual_annotation"]

    # Freeze header row
    ws.freeze_panes = "A2"

    # Autofilter
    ws.auto_filter.ref = ws.dimensions

    # Column widths
    widths = {
        "A": 12,  # pmid
        "B": 8,   # year
        "C": 32,  # journal
        "D": 55,  # title
        "E": 90,  # abstract
        "F": 18,  # manual_decision
        "G": 35,  # reason
        "H": 35,  # manual_surgical_problem
        "I": 35,  # manual_technical_solution
        "J": 40,  # augmentation_bridge
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Header style and wrapping
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="DDDDDD")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)

    # Wrap text for title/abstract/manual fields
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Set row height moderately, not huge
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 85

    # Data validation for manual_decision
    dv = DataValidation(
        type="list",
        formula1='"include,reject"',
        allow_blank=True
    )
    ws.add_data_validation(dv)
    dv.add(f"F2:F{ws.max_row}")

print("Done.")
print(f"CSV written:   {OUTPUT_CSV}")
print(f"Excel written: {OUTPUT_XLSX}")
print(f"Rows: {len(out)}")
print("Columns:")
print(list(out.columns))