from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ============================================================
# CONFIG
# ============================================================

BASE_DIR = Path(r"C:\Users\Max Stauffer\BreakingRules")

INPUT_CSV = (
    BASE_DIR
    / "openalex_llm_hypothesis_test_outputs"
    / "llm_test_50_random_hypotheses.csv"
)

OUT_XLSX = (
    BASE_DIR
    / "openalex_llm_hypothesis_test_outputs"
    / "llm_test_50_random_hypotheses_CLEAN.xlsx"
)


# ============================================================
# LOAD
# ============================================================

if not INPUT_CSV.exists():
    raise FileNotFoundError(INPUT_CSV)

df = pd.read_csv(INPUT_CSV, encoding="utf-8-sig", low_memory=False)

# Clean line breaks inside cells
for col in df.columns:
    if df[col].dtype == "object":
        df[col] = (
            df[col]
            .astype(str)
            .str.replace("\r", " ", regex=False)
            .str.replace("\n", " ", regex=False)
            .str.replace("\t", " ", regex=False)
            .str.replace(r"\s+", " ", regex=True)
            .str.strip()
        )

# Preferred column order
preferred_cols = [
    "llm_test_index",
    "review_rank",
    "final_rank",
    "annotated_rank",
    "ml_probability_top20_like",
    "concept_id",
    "concept_text",
    "problem_cluster_id",
    "problem_cluster_label",
    "mapped_solution_cluster_id",
    "solution_cluster_label",
    "hypothesis_title",
    "hypothesis_text",
    "mechanistic_rationale",
    "testing_route",
    "hypothesis_input_context",
    "error",
    "timestamp",
]

existing_preferred = [c for c in preferred_cols if c in df.columns]
remaining_cols = [c for c in df.columns if c not in existing_preferred]
df = df[existing_preferred + remaining_cols].copy()

# Review sheet: narrower and human-readable
review_cols = [
    "llm_test_index",
    "ml_probability_top20_like",
    "concept_text",
    "problem_cluster_label",
    "solution_cluster_label",
    "hypothesis_title",
    "hypothesis_text",
    "mechanistic_rationale",
    "testing_route",
    "problem_cluster_id",
    "mapped_solution_cluster_id",
    "concept_id",
]

review_cols = [c for c in review_cols if c in df.columns]
review_df = df[review_cols].copy()

# Add manual review columns
review_df.insert(0, "manual_keep", "")
review_df.insert(1, "manual_rating_1_to_5", "")
review_df.insert(2, "manual_comment", "")

# Summary sheet
summary_rows = [
    ["Input CSV", str(INPUT_CSV)],
    ["Output XLSX", str(OUT_XLSX)],
    ["Rows", len(df)],
    ["Unique OpenAlex concepts", df["concept_id"].nunique() if "concept_id" in df.columns else ""],
    ["Unique problem clusters", df["problem_cluster_id"].nunique() if "problem_cluster_id" in df.columns else ""],
    ["Unique solution clusters", df["mapped_solution_cluster_id"].nunique() if "mapped_solution_cluster_id" in df.columns else ""],
    ["Score definition", "ML probability only; no similarity term used in ranking/selection."],
]

summary_df = pd.DataFrame(summary_rows, columns=["Field", "Value"])


# ============================================================
# WRITE EXCEL
# ============================================================

with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
    review_df.to_excel(writer, sheet_name="Review_50", index=False)
    df.to_excel(writer, sheet_name="Full_Output", index=False)
    summary_df.to_excel(writer, sheet_name="Summary", index=False)


# ============================================================
# FORMAT EXCEL
# ============================================================

wb = load_workbook(OUT_XLSX)

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
subtle_fill = PatternFill("solid", fgColor="D9EAF7")
thin_gray = Side(style="thin", color="D9D9D9")
border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

text_cols_wide = {
    "concept_text",
    "problem_cluster_label",
    "solution_cluster_label",
    "hypothesis_title",
    "hypothesis_text",
    "mechanistic_rationale",
    "testing_route",
    "hypothesis_input_context",
    "manual_comment",
}

manual_cols = {
    "manual_keep",
    "manual_rating_1_to_5",
    "manual_comment",
}


def format_sheet(ws, table_name):
    max_row = ws.max_row
    max_col = ws.max_column

    # Header formatting
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Body formatting
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    # Manual columns styling
    headers = [ws.cell(row=1, column=i).value for i in range(1, max_col + 1)]
    for i, h in enumerate(headers, start=1):
        if h in manual_cols:
            for cell in ws.iter_cols(min_col=i, max_col=i, min_row=1, max_row=max_row):
                for c in cell:
                    c.fill = subtle_fill if c.row == 1 else PatternFill("solid", fgColor="F3F8FC")

    # Widths
    for i, header in enumerate(headers, start=1):
        col_letter = get_column_letter(i)

        if header in ["hypothesis_text", "mechanistic_rationale", "testing_route", "hypothesis_input_context"]:
            ws.column_dimensions[col_letter].width = 55
        elif header in ["concept_text", "problem_cluster_label", "solution_cluster_label"]:
            ws.column_dimensions[col_letter].width = 38
        elif header in ["hypothesis_title"]:
            ws.column_dimensions[col_letter].width = 42
        elif header in ["manual_comment"]:
            ws.column_dimensions[col_letter].width = 35
        elif header in ["manual_keep", "manual_rating_1_to_5"]:
            ws.column_dimensions[col_letter].width = 16
        elif header in ["ml_probability_top20_like"]:
            ws.column_dimensions[col_letter].width = 18
        elif header and "id" in str(header).lower():
            ws.column_dimensions[col_letter].width = 14
        else:
            ws.column_dimensions[col_letter].width = 18

    # Row heights
    ws.row_dimensions[1].height = 28
    for r in range(2, max_row + 1):
        ws.row_dimensions[r].height = 72 if ws.title == "Review_50" else 45

    # Freeze panes and autofilter/table
    ws.freeze_panes = "A2"

    if max_row >= 2 and max_col >= 1:
        table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
        tab = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

    # Number format
    for i, header in enumerate(headers, start=1):
        if header == "ml_probability_top20_like":
            for cell in ws.iter_cols(min_col=i, max_col=i, min_row=2, max_row=max_row):
                for c in cell:
                    c.number_format = "0.0000"


format_sheet(wb["Review_50"], "Review50Table")
format_sheet(wb["Full_Output"], "FullOutputTable")

# Summary formatting
ws = wb["Summary"]
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")
ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 120
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
ws.freeze_panes = "A2"

wb.save(OUT_XLSX)

print("DONE")
print("Clean Excel saved:")
print(OUT_XLSX)